import openpyxl.worksheet
from dataclasses import dataclass
from recording import Recording

# IDs of songs
SONG_IDS: dict[str, int] = {
    "Long Ago and Far Away": 0,
    "Summertime": 1,
    "My Little Suede Shoes": 2
}

# Names of songs
SONG_NAMES: dict[int, str] = {
    0: "Long Ago and Far Away",
    1: "Summertime",
    2: "My Little Suede Shoes",
}

# IDs of algorithms
ALGORITHM_IDS: dict[str, int] = {
    "Note Retrieval": 0,
    "Token Factor Oracle": 1,
    "Token Markov": 2,
}

# Names of algorithms
ALGORITHM_NAMES: dict[int, str] = {
    0: "Note Retrieval",
    1: "Token Factor Oracle",
    2: "Token Markov",
}

# Pitch classes of root, third and fifth of songs
SONG_PITCH_CLASSES: dict[str, list[int]] = {
    "Long Ago and Far Away": [7, 11, 2],
    "Summertime": [7, 10, 2],
    "My Little Suede Shoes": [3, 7, 10]
}

NUM_PUPILS = 5
NUM_PERFORMANCES = NUM_SONGS = NUM_ALGORITHMS = 3
NUM_SESSIONS = 2
NUM_QUESTIONS = 4

@dataclass
class PerformanceInfo:
    song: int
    algorithm: int
    recording: Recording
    responses: list[int]

@dataclass
class SessionInfo:
    performances: list[PerformanceInfo]

@dataclass
class PupilInfo:
    sessions: list[SessionInfo]


# (Retrieves pupil info from spreadsheet.xlsx)
def get_all_pupil_info(file_name: str = "spreadsheet.xlsx") -> list[PupilInfo]:
    # For every pupil, for every song, for every algorithm, there are 4 responses
    workbook: openpyxl.Workbook = openpyxl.load_workbook(file_name)

    res: list[PupilInfo] = []

    # For every pupil
    for pupil in range(NUM_PUPILS):
        pupil_info: PupilInfo = PupilInfo([])

        # Get Excel sheet, get starting row
        sheet = workbook[str(pupil + 1)]

        # For every session they do
        for session in range(NUM_SESSIONS):
            session_info: SessionInfo = SessionInfo([])

            # For every performance they do within that session
            for performance in range(NUM_PERFORMANCES):
                start_row: int = 8 * performance + 27 * session

                # Get performance info as text
                song_and_algorithm: str = str(sheet.cell(start_row + 2, 3).value)
                # Extract song and algorithm from text
                [song_str, algorithm_str] = song_and_algorithm.split(", ")
                song: int = SONG_IDS[song_str]
                algorithm: int = ALGORITHM_IDS[algorithm_str]

                file_path: str = f"recordings/{session + 1}/{pupil + 1}/{performance + 1}.notes"

                recording = None
                try:
                    recording: Recording = Recording(file_path)
                except:
                    pass

                responses: list[int] = [int(sheet.cell(start_row + 3 + i, 6).value) for i in range(NUM_QUESTIONS)]
                performance_info: PerformanceInfo = PerformanceInfo(song, algorithm, recording, responses)

                session_info.performances.append(performance_info)
            
            pupil_info.sessions.append(session_info)
        
        res.append(pupil_info)
    
    return res
